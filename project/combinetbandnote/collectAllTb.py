# -*- coding: UTF-8 -*-
import os
import sys
import shutil
import re
import xlrd
import pandas as pd
from openpyxl import load_workbook

# TODO:
# 汇总所有的TB
# 汇总所有的附注

# 合并策略：
# 1、不变
no_change = ["标准编码","科目余额表","本期ETY","新准则转换TB","新准则转换ETY","上期TB","上期ETY","审定报表分析性复核","重要性水平","资产表","负债表",
             "利润表","现金流量表","本期所有者权益变动表","上期所有者权益变动表","国有资本保值增值计算表","财务绩效评价指标",
             "校验","利息资本化校验","借款校验","税费校验表","薪酬校验表","折旧及摊销校验表","处置资产校验","政府补助校验",
             "递延所得税费用校验","投资收益核对","资产减值损失核对","子公司与少数股东现金流校验",
             "首次执行日前后金融资产分类和计量对比表","新旧金融工具调节表","金融资产减值准备调节表",
             "新金融工具准则对期初留存收益和其他综合收益的影响","新收入准则对期初财务报表的影响","新收入准则对期末资产负债表的影响",
             "新收入准则对利润表的影响","新租赁准则对期初报表的影响","最低经营租赁付款额与租赁负债调节表",
             "货币资金","受限制的货币资金","受限货币资金情况","交易性金融资产","以公允价值计量且其变动计入当期损益的金融资产",
             "衍生金融资产","应收票据分类原金融工具准则","应收票据分类新金融工具准则","已质押应收票据","已背书或贴现且在资产负债表日尚未到期的应收票据",
             "因出票人未履约而转为应收账款的票据","期末单项计提坏账准备的应收票据新金融工具准则","采用组合计提坏账准备的应收票据新金融工具准则",
             "应收票据坏账准备变动明细情况新金融工具准则","本期重要的应收票据坏账准备收回或转回情况新金融工具准则","本期实际核销的应收票据情况新金融工具准则",
             "应收账款期末数原金融工具准则","应收账款期初数原金融工具准则","应收账款期末数新金融工具准则","应收账款期初数新金融工具准则",
             "应收账款期末数首次新金融工具准则","应收账款期初数首次新金融工具准则","期末单项计提坏账准备的应收账款","采用账龄分析法计提坏账准备的应收账款原准则",
             "采用其他组合方法计提坏账准备的应收账款原准则","期末单项金额虽不重大但单项计提坏账准备的应收账款原准则","收回或转回的坏账准备情况",
             "本年实际核销的应收账款情况","按欠款方归集的年末余额前五名的应收账款情况",
             "采用组合计提坏账准备的应收账款首次执行","组合1名称首次执行","组合2名称首次执行","采用组合计提坏账准备的应收账款新金融工具","组合1名称新金融工具","组合2名称新金融工具",
             "应收账款坏账准备变动明细情况新金融工具准则","应收款项融资","应收款项融资已转让已背书或已贴现未到期","预付账款账龄明细",
             "账龄超过1年的大额预付款项情况","按欠款方归集的年末余额前五名的预付账款情况","其他应收款原准则","其他应收款期末数首次新金融工具准则",
             "其他应收款期初数首次新金融工具准则","其他应收款期末数新金融工具准则","其他应收款期初数新金融工具准则","应收利息分类","重要逾期利息",
             "应收股利明细","其他应收款项期末明细原准则","其他应收款项期初明细原准则","期末单项计提坏账准备的其他应收款","采用组合计提坏账准备的其他应收款首次执行",
             "采用组合计提坏账准备的其他应收款新金融工具准则","其他应收款账龄情况新金融工具准则","其他应收款坏账准备变动情况新金融工具准则",
             "采用账龄分析法计提坏账准备的其他应收款项原准则","采用其他组合方法计提坏账准备的其他应收款原准则","期末单项金额虽不重大但单项计提坏账准备的其他应收款原准则",
             "其他应收款收回或转回的坏账准备情况","本年实际核销的其他应收款情况","其他应收款按性质分类情况","按欠款方归集的年末金额前五名的其他应收款项情况",
             "存货明细情况","存货跌价准备明细情况",
             "合同履约成本","存货期末余额中借款费用资本化情况","存货成本倒闸表","合同资产情况","合同资产本期的重大变动","期末单项计提坏账准备的合同资产",
             "采用组合计提坏账准备的合同资产","持有待售资产的基本情况","持有待售资产减值准备情况","一年内到期的非流动资产","其他流动资产",
             "合同取得成本","债权投资","债权投资减值准备","期末重要的债权投资","可供出售金融资产情况","期末按公允价值计量的可供出售金融资产","可供出售权益工具严重下跌但未计提减值",
             "其他债权投资期末数","其他债权投资期初数","其他债权投资减值准备","期末重要的其他债权投资","持有至到期投资明细情况","期末重要的持有至到期投资",
             "长期应收款明细情况","长期应收款坏账准备变动情况新金融工具准则","因金融资产转移而终止确认的长期应收款","转移长期应收款且继续涉入形成的资产负债金额",
             "长期股权投资分类情况","长期股权投资子公司明细情况","长期股权投资合营企业明细情况","长期股权投资联营企业明细情况","对合营企业投资和联营企业投资国有企业",
             "其他权益工具投资明细","期末重要的其他权益工具投资国有企业","非交易性权益工具投资情况上市公司","其他非流动金融资产","采用成本计量模式的投资性房地产国有企业",
             "采用成本计量模式的投资性房地产上市公司","采用公允价值计量模式的投资性房地产","未办妥产权证书的投资性房地产金额及原因","固定资产汇总",
             "固定资产情况国有企业","固定资产情况上市公司","暂时闲置的固定资产情况","通过经营租赁租出的固定资产","未办妥产权证书的固定资产情况","固定资产清理",
             "在建工程汇总","在建工程情况","重要在建工程项目本期变动情况","本期计提在建工程减值准备情况","工程物资","生产性生物资产上市公司",
             "生产性生物资产国有企业","油气资产上市公司","油气资产国有企业","使用权资产上市公司","使用权资产国有企业","无形资产国有企业",
             "无形资产上市公司","未办妥产权证书的土地使用权情况","开发支出","商誉账面价值","商誉减值准备","长期待摊费用",
             "未经抵销的递延所得税资产","未经抵销的递延所得税负债","未确认递延所得税资产明细","未确认递延所得税资产的可抵扣亏损将于以下年度到期",
             "其他非流动资产","短期借款明细情况","已逾期未偿还的短期借款情况","应付账款","账龄超过一年的重要应付账款","预收款项","预收款项账龄表","账龄一年以上重要的预收款项",
             "应付职工薪酬明细情况","短期薪酬列示","设定提存计划列示","应交税费","应交增值税计提","其他应付款汇总","应付利息","重要的已逾期未支付的利息情况",
             "应付股利","账龄一年以上重要的应付股利","其他应付款项","账龄超过一年的重要其他应付款项","租赁负债","应付债券","应付债券的增减变动","长期借款","归属于权益工具持有者的信息",
             "长期应付款汇总","长期应付款","专项应付款","预计负债","递延收益","递延收益中政府补助项目","其他综合收益","盈余公积","未分配利润","营业收入与营业成本","主营业务收入与主营业务成本",
             "其他业务收入与其他业务成本","财务费用分类表","会计利润与所得税费用调整过程","当期所得税费用计算表","所得税项目计算","现金流量表补充资料",
             "现金流补充资料计算","股份支付总体情况","以权益结算的股份支付情况","以现金结算的股份支付情况","分部信息业务分部期末数",
             "分部信息业务分部期初数","本公司对主要客户的依赖程度","净资产收益率及每股收益上市公司","净资产收益率计算表","每股收益计算表",
             "分类表","信息分类表"



             ]

# 2、列合并，添加链接
columns_combine=[
    "本期TB","会计利润与所得税费用调整过程","本期支付的取得子公司的现金净额","现金流量表补充资料","本期收到的处置子公司的现金净额",
    "非经常性损益上市公司"]

# 列合并，合并所有列数据，待处理
#TODO：待处理，列索引不固定的处理
direct_combine_columns = [
"合并成本及商誉上市公司","被购买方于购买日可辨认资产负债上市公司", "合并成本上市公司","合并日被合并方资产负债的账面价值上市公司",
    "交易对于少数股东权益及归属于母公司所有者权益的影响上市公司","重要合营企业财务信息本期数上市公司","重要合营企业财务信息上期数上市公司",
    "重要联营企业财务信息本期数上市公司", "重要联营企业财务信息上期数上市公司", "不重要合营企业和联营企业的汇总信息上市公司",
]

#3、来源于母公司
parent = ["基础信息","实收资本","股本","母公司基本情况","关键管理人员薪酬"]

# 4、行合并，直接合并行数据
rows_combine = [
    "子企业情况-国有企业","表决权不足半数但能形成控制-国有企业","半数以上表决权但未控制-国有企业","本年不再纳入合并范围原子公司的情况-国有企业",
                "原子公司在处置日和上一会计期间资产负债表日的财务状况-国有企业","原子公司本年年初至处置日的经营成果-国有企业",
                "本年新纳入合并范围的主体-国有企业","本年发生的同一控制下企业合并情况-国有企业","本年发生的非同一控制下企业合并情况-国有企业",
                "本年发生的反向购买-国有企业","本年发生的吸收合并-国有企业",
                "货币资金明细表","受限货币资金明细表","交易性金融资产明细表","以公允价值计量且其变动计入当期损益的金融资产明细表",
                "衍生金融资产明细表","应收票据明细表","已质押票据明细表","已背书或贴现且在资产负债表日尚未到期的应收票据明细表",
                "因出票人未履约而转为应收账款的票据明细表","应收账款明细表","应收款项融资明细表","应收款项融资已转让已背书或已贴现未到期明细表",
                "预付账款明细表","应收利息明细表","应收股利明细表","其他应收款明细表","房地产开发成本","房地产开发产品","合同履约明细表",
                "确定可变现净值的具体依据","存货明细表","合同资产明细表","持有待售资产明细表","一年内到期的非流动资产明细表","其他流动资产明细表","合同取得成本明细表",
                "债权投资明细表","债权投资减值准备明细表","可供出售债务工具明细表","可供出售权益工具明细表","其他债权投资明细表","其他债权投资减值准备明细表",
                "长期应收款明细表","长期应收款减值准备明细表新金融工具","因金融资产转移而终止确认的长期应收款明细表","向投资企业转移资金的能力受到限制的有关情况国有企业",
                "长期股权投资明细表","合营企业和联营企业主要财务信息明细表","其他权益工具投资明细表","其他非流动金融资产明细表",
                "成本法核算投资性房地产明细表","采用公允价值计量模式的投资性房地产明细表","未办妥权证的投资性房地产明细表",
                "固定资产明细表","暂时闲置的固定资产明细表","经营租赁租出固定资产明细表","未办妥权证的固定资产明细表","固定资产清理明细表",
                "在建工程明细表","工程物资明细表","生产性生物资产明细表","油气资产明细表","使用权资产明细表","无形资产明细表",
                "未办妥产权证书的无形资产明细表","开发支出明细表","商誉明细表","长期待摊费用明细表","其他非流动资产明细表",
                "短期借款明细表","应付账款明细表","预收账款明细表","应付职工薪酬明细表","应交税费明细表","应付利息明细表","应付股利明细表",
                "其他应付款明细表","租赁负债明细表","长期借款明细表","长期应付款明细表","专项应付款明细表","应付债券明细表",
                "主营业务明细表","其他业务明细表","可抵扣暂时性差异明细表","应纳税暂时性差异明细表","可抵扣亏损",
                "对外担保明细表","其他或有事项明细表","预计负债明细表","递延收益明细表","未实现售后回租损益明细表",
                "非同一控制下企业合并上市公司","同一控制下企业合并上市公司","单次处置对子公司投资即丧失控制权上市公司","多次处置构成一揽子交易上市公司",
                "多次处置不构成一揽子交易上市公司","其他合并范围增加上市公司"," 其他合并范围减少上市公司","企业集团的构成上市公司",
                "重要的非全资子公司上市公司","重要非全资子企业期末资产负债上市公司","重要非全资子企业期初资产负债上市公司",
                "重要非全资子企业本期损益和现金流量情况上市公司","重要非全资子企业上期损益和现金流量情况上市公司",
                "在子公司的所有者权益份额发生变化的情况说明上市公司","重要的合营企业或联营企业上市公司","合营企业或联营企业发生的超额亏损上市公司",
                "重要的共同经营上市公司","第三层次公允价值计量项目期初与期末账面价值间的调节信息上市公司","不以公允价值计量的金融资产和金融负债的公允价值情况上市公司",
                ]

# 5、行合并，添加合并明细表
add_detail_combine=[
    "首次执行日前后金融资产分类和计量对比表","新旧金融工具调节表","金融资产减值准备调节表","新金融工具准则对期初留存收益和其他综合收益的影响",
    "新收入准则对期初财务报表的影响","新收入准则对期末资产负债表的影响","新收入准则对利润表的影响","新租赁准则对期初报表的影响",
    "最低经营租赁付款额与租赁负债调节表","主要税种及税率","由金融资产转移而终止确认的应收账款","转移应收账款且继续涉入形成的资产负债",
    "由金融资产转移而终止确认的其他应收款项","转移其他应收款且继续涉入形成的资产负债","按应收金额确认的政府补助",
    "持有至到期投资明细情况","期末重要的持有至到期投资","转移长期应收款且继续涉入形成的资产负债金额",
    "交易性金融负债","以公允价值计量且其变动计入当期损益的金融负债","衍生金融负债","应付票据","合同负债","持有待售负债",
    "一年内到期的非流动负债","其他流动负债","短期应付债券","期末发行在外的优先股永续债等金融工具情况",
    "发行在外的优先股永续债等金融工具变动情况","长期应付职工薪酬明细情况","设定受益计划义务现值",
    "计划资产","设定受益计划净负债","其他非流动负债","其他权益工具","资本公积","专项储备","税金及附加","销售费用","管理费用",
    "研发费用","财务费用","其他收益","投资收益","净敞口套期收益","公允价值变动损益","信用减值损失","资产减值损失","资产处置收益",
    "营业外收入","营业外支出","所得税费用","收到其他与经营活动有关的现金","支付其他与经营活动有关的现金","收到其他与投资活动有关的现金",
    "支付其他与投资活动有关的现金","收到其他与筹资活动有关的现金","支付其他与筹资活动有关的现金","现金及现金等价物的构成",
    "所有权或使用权受到限制的资产","外币货币性项目","其他关联方情况","采购商品接收劳务","出售商品提供劳务","本公司作为出租方","本公司作为承租方",
    "本公司作为承租方当期承担的租赁负债利息支出","本公司作为担保方","本公司作为担保方","本公司作为被担保方","关联方资金拆借","应收关联方款项",
    "应付关联方款项","关联方承诺上市公司","金融负债按剩余到期日分类期末数上市公司","金融负债按剩余到期日分类期初数上市公司","外币货币性项目上市公司",
    "外汇风险敏感性分析","利率敏感性分析上市公司","以公允价值计量的资产和负债的期末公允价值明细情况上市公司","资本承诺上市公司","按收入来源地划分的对外交易收入",
    "按资产所在地划分的非流动资产"
]

add_detail_combine_conditions={
    "首次执行日前后金融资产分类和计量对比表":("原金融工具准账面价值","新金融工具准账面价值"),
    "新旧金融工具调节表":("变更前报表","重分类","重新计量","变更后报表"),
    "金融资产减值准备调节表":("变更前报表","重分类","重新计量","变更后报表"),
    "新金融工具准则对期初留存收益和其他综合收益的影响":("未分配利润","盈余公积","其他综合收益"),
    "新收入准则对期初财务报表的影响":("变更前报表","变更后报表"),
    "新收入准则对期末资产负债表的影响":("新收入准则","旧收入准则"),
    "新收入准则对利润表的影响":("新收入准则","旧收入准则"),
    "新租赁准则对期初报表的影响":("变更前报表","变更后报表"),
    "最低经营租赁付款额与租赁负债调节表":("金额",),
    "主要税种及税率":(),
    "由金融资产转移而终止确认的应收账款":("终止确认金额",),
    "转移应收账款且继续涉入形成的资产负债":("期末数",),
    "由金融资产转移而终止确认的其他应收款项":("终止确认金额",),
    "转移其他应收款且继续涉入形成的资产负债":("期末数",),
    "按应收金额确认的政府补助":("期末余额",),
    "持有至到期投资明细情况":("期末余额","期初余额"),
    "期末重要的持有至到期投资":("面值",),
    "转移长期应收款且继续涉入形成的资产负债金额":("期末数",),
    "交易性金融负债":("期末数","期初数"),
    "以公允价值计量且其变动计入当期损益的金融负债":("期末数","期初数"),
    "衍生金融负债":("期末数","期初数"),
    "应付票据":("期末数","期初数"),
    "合同负债":("期末数","期初数"),
    "持有待售负债":("期末数","期初数"),
    "一年内到期的非流动负债":("期末数","期初数"),
    "其他流动负债":("期末数","期初数"),
    "短期应付债券":("面值",),
    "期末发行在外的优先股永续债等金融工具情况":("金额",),
    "发行在外的优先股永续债等金融工具变动情况":("期初账面价值","期末账面价值"),
    "长期应付职工薪酬明细情况":("期末数","期初数"),
    "设定受益计划义务现值":("本期数","上期数"),
    "计划资产":("本期数","上期数"),
    "设定受益计划净负债":("本期数","上期数"),
    "其他非流动负债":("期末数","期初数"),
    "其他权益工具":("期初账面价值","期末账面价值"),
    "资本公积":("期初数","期末数"),
    "专项储备":("期初数","期末数"),
    "税金及附加":("本期数","上年同期数"),
    "销售费用":("本期数","上年同期数"),
    "管理费用":("本期数","上年同期数"),
    "财务费用":("本期数","上年同期数"),
    "其他收益":("本期数","上年同期数"),
    "投资收益":("本期数","上年同期数"),
    "公允价值变动损益":("本期数","上年同期数"),
    "研发费用":("本期数","上年同期数"),
    "信用减值损失":("本期数","上年同期数"),
    "资产减值损失":("本期数","上年同期数"),
    "资产处置收益":("本期数","上年同期数"),
    "净敞口套期收益":("本期数","上年同期数"),
    "营业外收入":("本期数","上年同期数"),
    "营业外支出":("本期数","上年同期数"),
    "所得税费用":("本期数","上年同期数"),
    "收到其他与经营活动有关的现金":("本期数","上年同期数"),
    "支付其他与经营活动有关的现金":("本期数","上年同期数"),
    "收到其他与投资活动有关的现金":("本期数","上年同期数"),
    "支付其他与投资活动有关的现金":("本期数","上年同期数"),
    "收到其他与筹资活动有关的现金":("本期数","上年同期数"),
    "支付其他与筹资活动有关的现金":("本期数","上年同期数"),
    "现金及现金等价物的构成":("期末数","期初数"),
    "所有权或使用权受到限制的资产":("期末账面价值",),
    "外币货币性项目":("期末折算成人民币余额",),
    "其他关联方情况":(),
    "采购商品接收劳务":("本期数","上年同期数"),
    "出售商品提供劳务":("本期数","上年同期数"),
    "本公司作为出租方":("本期数","上年同期数"),
    "本公司作为承租方":("本期数","上年同期数"),
    "本公司作为承租方当期承担的租赁负债利息支出":("本期数","上年同期数"),
    "本公司作为担保方":("担保金额",),
    "本公司作为被担保方":("担保金额",),
    "关联方资金拆借":("拆借金额",),
    "应收关联方款项":("期末账面余额","期初账面余额"),
    "应付关联方款项":("期末数","期初数"),
    "关联方承诺上市公司":("期末数","期初数"),
    "金融负债按剩余到期日分类期末数上市公司":("账面价值",),
    "金融负债按剩余到期日分类期初数上市公司":("账面价值",),
    "外币货币性项目上市公司":("期末数","期初数"),
    "外汇风险敏感性分析":("对本期净利润的影响",),
    "利率敏感性分析上市公司":("对本期净利润的影响",),
    "以公允价值计量的资产和负债的期末公允价值明细情况上市公司":("合计",),
    "资本承诺上市公司":("期末数","期初数"),
    "按收入来源地划分的对外交易收入":("本期数","上期数"),
    "按资产所在地划分的非流动资产":("本期数","上期数"),
}

# 检查是否有遗漏的表格
def check_is_not_exist():
    all_tables = [*no_change,*columns_combine,*parent,*rows_combine,*add_detail_combine,*direct_combine_columns]

    import xlrd

    file = "D:/auditReport/project/nationalmodel.xlsx"
    excels = xlrd.open_workbook(file)
    sheet_names = excels.sheet_names()
    for sheet_name in sheet_names:
        if sheet_name not in all_tables:
            print(sheet_name)

# 检查每个表格的表头是否有重复的标题
def check_title_is_duplicated():
    import xlrd
    file = "D:/auditReport/project/nationalmodel.xlsx"
    bk=xlrd.open_workbook(file)
    excels = xlrd.open_workbook(file)
    sheet_names = excels.sheet_names()
    for sheet_name in sheet_names:
        sh = bk.sheet_by_name(sheet_name)
        row_data = sh.row_values(0)
        if len(row_data)!=len(set(row_data)):
            print(sheet_name,row_data)

# 复制一列到新的sheet中,返回新的sheet
def copy_column_index(wb,old_sheet_name,new_sheet_name,old_sheet_column_num,new_sheet_column_num):
    sheetnames = wb.sheetnames
    if old_sheet_name not in sheetnames:
        raise Exception("无法在model中找到{}".format(old_sheet_name))
    if new_sheet_name in sheetnames:
        raise Exception("model中已经存在{}".format(new_sheet_name))

    ws1 = wb[old_sheet_name]
    ws1_maxrows = ws1.max_row
    items = []
    i = 1
    while i < ws1_maxrows:
        items.append(ws1.cell(i, old_sheet_column_num).value)
        i = i + 1

    num = sheetnames.index(old_sheet_name)
    ws2 = wb.create_sheet(new_sheet_name, num)
    i = 1
    while i < ws1_maxrows:
        ws2.cell(i, new_sheet_column_num).value = items[i - 1]
        i = i + 1

    return wb,ws1,ws2,ws1_maxrows


# 计算excel中的列
def get_column_alpha(num):
    basics = ["a","b","c","d","e","f","g","h","i",
              "j","k","l","m","n","o","p","q","r",
              "s","t","u","v","w","x","y","z"
              ]
    if num <= len(basics):
        return basics[num-1]
    else:
        first = int(num / 26)
        if first>26:
            raise Exception("can't get column alpha num")
        last = num%26
        if last==0:
            first = first-1
            last = 26
        first_alpha = basics[first-1]
        last_alpha = basics[last-1]
        return "{}{}".format(first_alpha,last_alpha)


# 列合并，添加链接
def combine_columns_add_link(wb,all_files,contrast_names):
    for combine_column_item in contrast_names:
        # 添加合并表，并复制索引列
        wb, single_sheet, combine_sheet, max_rows = copy_column_index(wb, combine_column_item["name"], combine_column_item["combinename"], 1, 1)
        add_column_link(all_files, max_rows, combine_sheet,  combine_column_item["name"],  combine_column_item["column"], combine_column_item["addtotal"])
    return wb

# 删除空白列
def delete_na_columns(df):
    # 进行统计非空为0
    col = df.count() == 0  # 返回bool数组
    for i in range(len(col)):
        if col[i]:
            df.drop(labels=col.index[i], axis=1, inplace=True)
    return df

# 列合并，直接合并数据：
# TODO:待处理
def combine_columns_direct():
    path = 'D:/auditReport/project/combinetbandnote/tbs'
    allow_suffix = ".xlsx"
    all_files = [f for f in os.listdir(path) if f.endswith(allow_suffix)]
    # 合并工作表
    combine_excel = "nationalmodel.xlsx"
    combine_path = os.path.join(path, combine_excel)
    # 获取所有明细表并合并
    dfs = []
    for combine_sheet_name in direct_combine_columns:
        for file in all_files:
            if file.startswith(combine_excel):
                continue
            filepath = os.path.join(path, file)
            df = pd.read_excel(filepath, sheet_name=combine_sheet_name)
            df = delete_na_columns(df)
            if len(df.columns) > 1:
                dfs.append(df)
        if len(dfs) > 0:
            df_res = pd.concat(dfs,axis=1)
            #删除重复列
            df_res = df_res.T.drop_duplicates().T
            book = load_workbook(combine_path)
            writer = pd.ExcelWriter(combine_path, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_res.to_excel(writer, combine_sheet_name, index=False)
            writer.save()
            dfs = []


# 行合并，直接合并数据
def combine_rows(path,combine_excel):

    allow_suffix = ".xlsx"
    all_files = [f for f in os.listdir(path) if f.endswith(allow_suffix)]
    # 合并工作表
    combine_path = os.path.join(path, combine_excel)
    wb = load_workbook(combine_path)
    # 获取所有明细表并合并
    dfs = []
    for combine_sheet_name in rows_combine:
        if wb[combine_sheet_name].sheet_state=="hidden":
            continue
        for file in all_files:
            if file.startswith(combine_excel):
                continue
            filepath = os.path.join(path, file)
            df = pd.read_excel(filepath, sheet_name=combine_sheet_name)
            df = df[df.iloc[:, 0].notna()]
            if len(df)>0:
                dfs.append(df)
        if len(dfs)>0:
            df_res = pd.concat(dfs, ignore_index=True)

            book = load_workbook(combine_path)
            writer = pd.ExcelWriter(combine_path, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_res.to_excel(writer, combine_sheet_name, index=False)
            writer.save()
            dfs=[]

# 获取单位信息
def get_company_name(wb):
    ws = wb["基础信息"]
    return ws.cell(1, 2).value

#

# 添加明细表和单位简称，创建合并明细表
def add_detail_combine_table(wb,old_sheet_name):
    new_sheet_name = "{}明细表".format(old_sheet_name)
    sheetnames = wb.sheetnames
    if new_sheet_name in sheetnames:
        return
    num = sheetnames.index(old_sheet_name)
    wb.create_sheet(new_sheet_name, num+1)

def create_all_combine_tables(wb):
    for name in add_detail_combine:
        if wb[name].sheet_state=="hidden":
            continue
        add_detail_combine_table(wb,name)

# 根据条件筛选
def filter_df(df,conditions):
    if len(conditions)==0:
        return df
    s = False
    for condition in conditions:
        try:
            s = s | (df[condition].abs() > 0)
        except Exception as e:
            print(e)
            return df
    df1 = df[s]
    return df1

# 添加明细表进行合并
def add_detail_combine_rows(path,combine_excel):
    allow_suffix = ".xlsx"
    all_files = [f for f in os.listdir(path) if f.endswith(allow_suffix)]
    # 合并工作表
    combine_path = os.path.join(path, combine_excel)
    wb = load_workbook(combine_path)

    # 创建合并明细表
    create_all_combine_tables(wb)
    wb.save(combine_path)
    # 获取所有明细表并合并
    dfs = []
    for sheet_name in add_detail_combine:
        if wb[sheet_name].sheet_state=="hidden":
            continue
        for file in all_files:
            if file.startswith(combine_excel):
                continue
            filepath = os.path.join(path, file)
            # 获取单位名称
            newwb = load_workbook(filepath)
            company_name = get_company_name(newwb)
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            df = filter_df(df,add_detail_combine_conditions[sheet_name])
            if len(df)>0:
                df.insert(0, "单位简称",company_name)
                dfs.append(df)
        if len(dfs)>0:
            df_res = pd.concat(dfs, ignore_index=True)
            book = load_workbook(combine_path)
            writer = pd.ExcelWriter(combine_path, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_res.to_excel(writer, "{}明细表".format(sheet_name), index=False)
            writer.save()
            dfs=[]



# 添加合并链接,并添加合计
def add_column_link(all_files,max_rows,combine_sheet,sheetname,combine_column,is_add_total):

    for key,filename  in enumerate(all_files):
        i = 1
        while i < max_rows:
            if i==1:
                combine_sheet.cell(i,key+2).value = filename.replace(".xlsx","")
            else:
                combine_sheet.cell(i,key+2).value="='[{}]{}'!${}${}".format(filename,sheetname,combine_column,i)
            i = i + 1
    if is_add_total:
        i = 1
        while i < max_rows:
            if i == 1:
                combine_sheet.cell(i, len(all_files) + 2).value = "合计"
            else:
                combine_sheet.cell(i, len(all_files) + 2).value = "=SUM(B{}:{}{})".format(i, get_column_alpha(
                    len(all_files) + 1), i)
            i = i + 1


# 获取改表格的列数
def get_table_column_num(path,sheetname):
    wb = load_workbook(path)
    ws = wb[sheetname]
    return ws.max_column

# 获取链接列配置表
def get_row_link_setting(path):
    df = pd.read_excel(path, sheet_name="setting")
    records = df.to_dict(orient='records')
    return records

# 获取某表格配置的最大链接行数
def get_table_max_link_row_num(sheetname,settings):
    for setting in settings:
        if setting["sheetname"]==sheetname:
            return setting["link_column_num"]
    return 100

# 将行合并也改为链接方式
def add_row_combine_link(path,combine_excel):
    '''

    :param path: 合并文件夹
    :param combine_excel: 合并excel
    :return:
    '''
    allow_suffix = ".xlsx"
    all_files = [f for f in os.listdir(path) if (f.endswith(allow_suffix) and not f.startswith(combine_excel) and not f.startswith("~$"))]
    target = os.path.join(path, combine_excel)
    settings = get_row_link_setting(target)
    wb = load_workbook(target)

    for sheetname in rows_combine:
        print(sheetname)
        combine_sheet = wb[sheetname]
        link_rows_num = get_table_max_link_row_num(sheetname, settings)
        column_num = get_table_column_num(target, sheetname)
        for key, filename in enumerate(all_files):
            i = 2
            while i < link_rows_num:
                j=1
                while j<column_num+1:
                    combine_column = get_column_alpha(j)
                    combine_sheet.cell(i+key*link_rows_num,j).value = "='[{}]{}'!${}${}".format(filename,sheetname,combine_column,i)
                    j = j+1
                i = i + 1
    wb.save(target)

# 替换TB和数字
def filter_filename(filename):
    filename = filename.replace("TB","")
    filename = filename.replace(".xlsx","")
    pattern = re.compile("\d+")
    res = pattern.sub("",filename)
    return res

# 将增加单位名称统计也改为link
def add_detail_link(path,combine_excel):
    allow_suffix = ".xlsx"
    all_files = [f for f in os.listdir(path) if
                 (f.endswith(allow_suffix) and not f.startswith(combine_excel) and not f.startswith("~$"))]
    # 合并工作表
    combine_path = os.path.join(path, combine_excel)
    settings = get_row_link_setting(combine_path)
    wb = load_workbook(combine_path)

    # 创建合并明细表
    create_all_combine_tables(wb)
    wb.save(combine_path)
    # 获取所有明细表并合并
    for sheetname in add_detail_combine:
        print(sheetname)
        if wb[sheetname].sheet_state=="hidden":
            continue
        # 获取最大行和最大列
        column_num = wb[sheetname].max_column
        link_rows_num = get_table_max_link_row_num(sheetname, settings)
        # 合并明细表
        combine_sheet_name = "{}明细表".format(sheetname)
        combine_sheet = wb[combine_sheet_name]
        # 添加标题
        t = 0
        while t < column_num+1:
            if t==0:
                combine_sheet.cell(1, t+1).value = "单位简称"
            else:
                combine_sheet.cell(1, t + 1).value = wb[sheetname].cell(1,t).value
            t = t+1

        # 添加内容
        for key, filename in enumerate(all_files):
            companyname = filter_filename(filename)
            i = 2
            while i < link_rows_num:
                j = 1
                while j <= column_num + 1:
                    if j==1:
                        combine_sheet.cell(i + key * link_rows_num, j).value = companyname
                    else:
                        combine_column = get_column_alpha(j-1)
                        combine_sheet.cell(i + key * link_rows_num, j).value = "='[{}]{}'!${}${}".format(filename,
                                                                                                         sheetname,
                                                                                                     combine_column, i)
                    j = j + 1
                i = i + 1
    wb.save(combine_path)

# 添加列合并链接，不能增减行
def add_link_columns_combine(source,path):
    # 合并工作表
    allow_suffix = ".xlsx"
    combine_excel = os.path.basename(source)
    all_files = [f for f in os.listdir(path) if
                 (f.endswith(allow_suffix) and not f.startswith(combine_excel) and not f.startswith("~$"))]

    # 复制文件到合并文件夹中
    try:
        shutil.copy(source, path)
    except IOError as e:
        print("Unable to copy file. %s" % e)
    except:
        print("Unexpected error:", sys.exc_info())
    # adding exception handling
    target = os.path.join(path,combine_excel)
    wb = load_workbook(target)
    contrast_names = [
        {"name": "本期TB", "combinename": "合并TB", "column": "H", "addtotal": True},
        {"name": "现金流量表", "combinename": "合并现金流量表", "column": "B", "addtotal": True},
        {"name": "会计利润与所得税费用调整过程", "combinename": "合并会计利润与所得税费用调整过程", "column": "B", "addtotal": True},
        {"name": "本期支付的取得子公司的现金净额", "combinename": "合并本期支付的取得子公司的现金净额", "column": "B", "addtotal": True},
        {"name": "将净利润调节为经营活动现金流量", "combinename": "合并将净利润调节为经营活动现金流量", "column": "B", "addtotal": True},
        {"name": "本期收到的处置子公司的现金净额", "combinename": "合并本期收到的处置子公司的现金净额", "column": "B", "addtotal": True},
        {"name": "非经常性损益上市公司", "combinename": "合并非经常性损益上市公司", "column": "B", "addtotal": True},
        {"name": "现金及现金等价物的构成", "combinename": "合并现金及现金等价物的构成", "column": "B", "addtotal": True},
    ]
    wb = combine_columns_add_link(wb, all_files, contrast_names)

    wb.save(target)

def test_add_detail():
    for item in add_detail_combine:
        if item not in add_detail_combine_conditions:
            print(item)

if __name__ == '__main__':
    # check_is_not_exist()
    # model = r'D:\auditReport\project\model.xlsx'
    # combinepath = r'D:\我的文件2022\杭州高新技术产业开发区资产经营有限公司2021年三季报\杭州高新区三季报\高新区资产经营2021年三季报TB\高新区资产经营本期TB -2020年'
    # combine_excel = "台州恒金创业投资有限公司合并.xlsx"
    # 1、复制模板，并添加列合并链接
    # add_link_columns_combine(model,combinepath)
    path = r"D:\我的文件2022\杭州高新技术产业开发区资产经营有限公司2021年三季报\杭州高新区三季报\高新区资产经营2021年三季报TB\高新区资产经营本期TB -2020年"
    combine_excel = "杭州高新技术产业开发区资产经营有限公司合并TB.xlsx"
    # 2、添加行链接
    add_row_combine_link(path, combine_excel)
    # 3、添加明细表链接
    add_detail_link(path, combine_excel)

