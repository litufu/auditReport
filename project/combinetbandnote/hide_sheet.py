from openpyxl import load_workbook

tables = [
    {"name": "基础信息", "condition": "all"},
    {"name": "标准编码", "condition": "all"},
    {"name": "科目余额表", "condition": "all"},
    {"name": "本期TB", "condition": "all"},
    {"name": "本期ETY", "condition": "all"},
    {"name": "新准则转换TB", "condition": "all"},
    {"name": "新准则转换ETY", "condition": "all"},
    {"name": "上期TB", "condition": "all"},
    {"name": "上期ETY", "condition": "all"},
    {"name": "审定报表分析性复核", "condition": "all"},
    {"name": "重要性水平", "condition": "all"},
    {"name": "资产表", "condition": "all"},
    {"name": "负债表", "condition": "all"},
    {"name": "利润表", "condition": "all"},
    {"name": "现金流量表", "condition": "all"},
    {"name": "本期所有者权益变动表", "condition": "all"},
    {"name": "上期所有者权益变动表", "condition": "all"},
    {"name": "国有资本保值增值计算表", "condition": "all"},
    {"name": "财务绩效评价指标", "condition": "all"},
    {"name": "校验", "condition": "all"},
    {"name": "利息资本化校验", "condition": "all"},
    {"name": "借款校验", "condition": "all"},
    {"name": "税费校验表", "condition": "all"},
    {"name": "薪酬校验表", "condition": "all"},
    {"name": "折旧及摊销校验表", "condition": "all"},
    {"name": "处置资产校验", "condition": "all"},
    {"name": "政府补助校验", "condition": "all"},
    {"name": "递延所得税费用校验", "condition": "all"},
    {"name": "投资收益核对", "condition": "all"},
    {"name": "资产减值损失核对", "condition": "all"},
    {"name": "子公司与少数股东现金流校验", "condition": "all"},
    {"name": "首次执行日前后金融资产分类和计量对比表", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "新旧金融工具调节表", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "金融资产减值准备调节表", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "新金融工具准则对期初留存收益和其他综合收益的影响", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "新收入准则对期初财务报表的影响", "condition": "firstTimeNewIncomeCriteria"},
    {"name": "新收入准则对期末资产负债表的影响", "condition": "firstTimeNewIncomeCriteria"},
    {"name": "新收入准则对利润表的影响", "condition": "firstTimeNewIncomeCriteria"},
    {"name": "新租赁准则对期初报表的影响", "condition": "firstTimeNewLeaseCriteria"},
    {"name": "最低经营租赁付款额与租赁负债调节表", "condition": "firstTimeNewLeaseCriteria"},
    {"name": "主要税种及税率", "condition": "all"},
    {"name": "子企业情况-国有企业", "condition": "national"},
    {"name": "表决权不足半数但能形成控制-国有企业", "condition": "national"},
    {"name": "半数以上表决权但未控制-国有企业", "condition": "national"},
    {"name": "本年不再纳入合并范围原子公司的情况-国有企业", "condition": "national"},
    {"name": "原子公司在处置日和上一会计期间资产负债表日的财务状况-国有企业", "condition": "national"},
    {"name": "原子公司本年年初至处置日的经营成果-国有企业", "condition": "national"},
    {"name": "本年新纳入合并范围的主体-国有企业", "condition": "national"},
    {"name": "本年发生的同一控制下企业合并情况-国有企业", "condition": "national"},
    {"name": "本年发生的非同一控制下企业合并情况-国有企业", "condition": "national"},
    {"name": "本年发生的反向购买-国有企业", "condition": "national"},
    {"name": "本年发生的吸收合并-国有企业", "condition": "national"},
    {"name": "货币资金", "condition": "all"},
    {"name": "受限制的货币资金", "condition": "all"},
    {"name": "受限货币资金情况", "condition": "all"},
    {"name": "货币资金明细表", "condition": "all"},
    {"name": "受限货币资金明细表", "condition": "all"},
    {"name": "交易性金融资产", "condition":("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "交易性金融资产明细表", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "以公允价值计量且其变动计入当期损益的金融资产", "condition": ("oldFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "以公允价值计量且其变动计入当期损益的金融资产明细表", "condition": ("oldFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "衍生金融资产", "condition": "all"},
    {"name": "衍生金融资产明细表", "condition": "all"},
    {"name": "应收票据分类原金融工具准则", "condition": "oldFinancialInstruments"},
    {"name": "应收票据分类新金融工具准则", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "已质押应收票据", "condition": "all"},
    {"name": "已质押票据明细表", "condition": "all"},
    {"name": "已背书或贴现且在资产负债表日尚未到期的应收票据", "condition": "all"},
    {"name": "已背书或贴现且在资产负债表日尚未到期的应收票据明细表", "condition": "all"},
    {"name": "因出票人未履约而转为应收账款的票据", "condition": "all"},
    {"name": "因出票人未履约而转为应收账款的票据明细表", "condition": "all"},
    {"name": "期末单项计提坏账准备的应收票据新金融工具准则", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "采用组合计提坏账准备的应收票据新金融工具准则", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "应收票据坏账准备变动明细情况新金融工具准则", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "本期重要的应收票据坏账准备收回或转回情况新金融工具准则", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "本期实际核销的应收票据情况新金融工具准则", "condition":("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "应收票据明细表", "condition": "all"},
    {"name": "应收账款期末数原金融工具准则", "condition": "oldFinancialInstruments"},
    {"name": "应收账款期初数原金融工具准则", "condition": "oldFinancialInstruments"},
    {"name": "应收账款期末数新金融工具准则", "condition": "newFinancialInstruments"},
    {"name": "应收账款期初数新金融工具准则", "condition": "newFinancialInstruments"},
    {"name": "应收账款期末数首次新金融工具准则", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "应收账款期初数首次新金融工具准则", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "期末单项计提坏账准备的应收账款", "condition": "all"},
    {"name": "采用账龄分析法计提坏账准备的应收账款原准则", "condition": "oldFinancialInstruments"},
    {"name": "采用其他组合方法计提坏账准备的应收账款原准则", "condition": "oldFinancialInstruments"},
    {"name": "期末单项金额虽不重大但单项计提坏账准备的应收账款原准则", "condition": "oldFinancialInstruments"},
    {"name": "收回或转回的坏账准备情况", "condition": "all"},
    {"name": "本年实际核销的应收账款情况", "condition": "all"},
    {"name": "按欠款方归集的年末余额前五名的应收账款情况", "condition": "all"},
    {"name": "由金融资产转移而终止确认的应收账款", "condition": "all"},
    {"name": "转移应收账款且继续涉入形成的资产负债", "condition": "all"},
    {"name": "采用组合计提坏账准备的应收账款首次执行", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "组合1名称首次执行", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "组合2名称首次执行", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "采用组合计提坏账准备的应收账款新金融工具", "condition": "newFinancialInstruments"},
    {"name": "组合1名称新金融工具", "condition": "newFinancialInstruments"},
    {"name": "组合2名称新金融工具", "condition": "newFinancialInstruments"},
    {"name": "应收账款坏账准备变动明细情况新金融工具准则", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "应收账款明细表", "condition": "all"},
    {"name": "应收款项融资", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "应收款项融资明细表", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "应收款项融资已转让已背书或已贴现未到期", "condition":("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "应收款项融资已转让已背书或已贴现未到期明细表", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "预付账款账龄明细", "condition": "all"},
    {"name": "账龄超过1年的大额预付款项情况", "condition": "all"},
    {"name": "按欠款方归集的年末余额前五名的预付账款情况", "condition": "all"},
    {"name": "预付账款明细表", "condition": "all"},
    {"name": "其他应收款原准则", "condition": "oldFinancialInstruments"},
    {"name": "其他应收款期末数首次新金融工具准则", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "其他应收款期初数首次新金融工具准则", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "其他应收款期末数新金融工具准则", "condition": "newFinancialInstruments"},
    {"name": "其他应收款期初数新金融工具准则", "condition": "newFinancialInstruments"},
    {"name": "应收利息分类", "condition": "all"},
    {"name": "重要逾期利息", "condition": "all"},
    {"name": "应收利息明细表", "condition": "all"},
    {"name": "应收股利明细", "condition": "all"},
    {"name": "应收股利明细表", "condition": "all"},
    {"name": "其他应收款项期末明细原准则", "condition": "oldFinancialInstruments"},
    {"name": "其他应收款项期初明细原准则", "condition": "oldFinancialInstruments"},
    {"name": "期末单项计提坏账准备的其他应收款", "condition": "all"},
    {"name": "采用组合计提坏账准备的其他应收款首次执行", "condition": "firstTimeNewFinancialInstruments"},
    {"name": "采用组合计提坏账准备的其他应收款新金融工具准则", "condition": "newFinancialInstruments"},
    {"name": "其他应收款账龄情况新金融工具准则", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "其他应收款坏账准备变动情况新金融工具准则", "condition":("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "其他应收款减值准备明细表新金融工具准则", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "采用账龄分析法计提坏账准备的其他应收款项原准则", "condition": "oldFinancialInstruments"},
    {"name": "采用其他组合方法计提坏账准备的其他应收款原准则", "condition": "oldFinancialInstruments"},
    {"name": "期末单项金额虽不重大但单项计提坏账准备的其他应收款原准则", "condition": "oldFinancialInstruments"},
    {"name": "其他应收款收回或转回的坏账准备情况", "condition": "all"},
    {"name": "本年实际核销的其他应收款情况", "condition": "all"},
    {"name": "其他应收款按性质分类情况", "condition": "all"},
    {"name": "按欠款方归集的年末金额前五名的其他应收款项情况", "condition": "all"},
    {"name": "按应收金额确认的政府补助", "condition": "all"},
    {"name": "由金融资产转移而终止确认的其他应收款项", "condition": "all"},
    {"name": "转移其他应收款且继续涉入形成的资产负债", "condition": "all"},
    {"name": "其他应收款明细表", "condition": "all"},
    {"name": "存货明细情况", "condition": "all"},
    {"name": "房地产开发成本", "condition": "all"},
    {"name": "房地产开发产品", "condition": "all"},
    {"name": "合同履约成本", "condition": "newIncomeCriteria"},
    {"name": "合同履约明细表", "condition": "newIncomeCriteria"},
    {"name": "存货跌价准备明细情况", "condition": "all"},
    {"name": "确定可变现净值的具体依据", "condition": "all"},
    {"name": "存货期末余额中借款费用资本化情况", "condition": "all"},
    {"name": "存货明细表", "condition": "all"},
    {"name": "存货成本倒闸表", "condition": "all"},
    {"name": "合同资产情况", "condition": "newIncomeCriteria"},
    {"name": "合同资产本期的重大变动", "condition": "newIncomeCriteria"},
    {"name": "期末单项计提坏账准备的合同资产", "condition": "newIncomeCriteria"},
    {"name": "采用组合计提坏账准备的合同资产", "condition": "newIncomeCriteria"},
    {"name": "合同资产明细表", "condition": "newIncomeCriteria"},
    {"name": "持有待售资产的基本情况", "condition": "all"},
    {"name": "持有待售资产减值准备情况", "condition": "all"},
    {"name": "持有待售资产明细表", "condition": "all"},
    {"name": "一年内到期的非流动资产", "condition": "all"},
    {"name": "一年内到期的非流动资产明细表", "condition": "all"},
    {"name": "其他流动资产", "condition": "all"},
    {"name": "其他流动资产明细表", "condition": "all"},
    {"name": "合同取得成本", "condition": "newIncomeCriteria"},
    {"name": "合同取得成本明细表", "condition": "newIncomeCriteria"},
    {"name": "债权投资", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "债权投资明细表", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "债权投资减值准备", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "债权投资减值准备明细表", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "期末重要的债权投资", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "可供出售金融资产情况", "condition": "oldFinancialInstruments"},
    {"name": "期末按公允价值计量的可供出售金融资产", "condition": "oldFinancialInstruments"},
    {"name": "可供出售权益工具严重下跌但未计提减值", "condition": "oldFinancialInstruments"},
    {"name": "可供出售债务工具明细表", "condition": "oldFinancialInstruments"},
    {"name": "可供出售权益工具明细表", "condition": "oldFinancialInstruments"},
    {"name": "其他债权投资期末数", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "其他债权投资期初数", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "其他债权投资明细表", "condition":("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "其他债权投资减值准备", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "其他债权投资减值准备明细表", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "期末重要的其他债权投资", "condition":("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "持有至到期投资明细情况", "condition": "oldFinancialInstruments"},
    {"name": "期末重要的持有至到期投资", "condition": "oldFinancialInstruments"},
    {"name": "长期应收款明细情况", "condition": "all"},
    {"name": "长期应收款明细表", "condition": "all"},
    {"name": "长期应收款坏账准备变动情况新金融工具准则", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "长期应收款减值准备明细表新金融工具", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "因金融资产转移而终止确认的长期应收款", "condition": "all"},
    {"name": "因金融资产转移而终止确认的长期应收款明细表", "condition": "all"},
    {"name": "转移长期应收款且继续涉入形成的资产负债金额", "condition": "all"},
    {"name": "长期股权投资分类情况", "condition": "all"},
    {"name": "长期股权投资子公司明细情况", "condition": "all"},
    {"name": "长期股权投资合营企业明细情况", "condition": "all"},
    {"name": "长期股权投资联营企业明细情况", "condition": "all"},
    {"name": "对合营企业投资和联营企业投资国有企业", "condition": "national"},
    {"name": "向投资企业转移资金的能力受到限制的有关情况国有企业", "condition": "national"},
    {"name": "长期股权投资明细表", "condition": "all"},
    {"name": "合营企业和联营企业主要财务信息明细表", "condition": "all"},
    {"name": "其他权益工具投资明细", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "期末重要的其他权益工具投资国有企业", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "非交易性权益工具投资情况上市公司", "condition":("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "其他权益工具投资明细表", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "其他非流动金融资产", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "其他非流动金融资产明细表", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "采用成本计量模式的投资性房地产国有企业", "condition": "national"},
    {"name": "采用成本计量模式的投资性房地产上市公司", "condition": "list"},
    {"name": "成本法核算投资性房地产明细表", "condition": "all"},
    {"name": "采用公允价值计量模式的投资性房地产", "condition": "all"},
    {"name": "采用公允价值计量模式的投资性房地产明细表", "condition": "all"},
    {"name": "未办妥产权证书的投资性房地产金额及原因", "condition": "all"},
    {"name": "未办妥权证的投资性房地产明细表", "condition": "all"},
    {"name": "固定资产汇总", "condition": "all"},
    {"name": "固定资产情况国有企业", "condition": "national"},
    {"name": "固定资产情况上市公司", "condition": "list"},
    {"name": "固定资产明细表", "condition": "all"},
    {"name": "暂时闲置的固定资产情况", "condition": "all"},
    {"name": "暂时闲置的固定资产明细表", "condition": "all"},
    {"name": "通过经营租赁租出的固定资产", "condition": "all"},
    {"name": "经营租赁租出固定资产明细表", "condition": "all"},
    {"name": "未办妥产权证书的固定资产情况", "condition": "all"},
    {"name": "未办妥权证的固定资产明细表", "condition": "all"},
    {"name": "固定资产清理", "condition": "all"},
    {"name": "固定资产清理明细表", "condition": "all"},
    {"name": "在建工程汇总", "condition": "all"},
    {"name": "在建工程情况", "condition": "all"},
    {"name": "重要在建工程项目本期变动情况", "condition": "all"},
    {"name": "本期计提在建工程减值准备情况", "condition": "all"},
    {"name": "在建工程明细表", "condition": "all"},
    {"name": "工程物资", "condition": "all"},
    {"name": "工程物资明细表", "condition": "all"},
    {"name": "生产性生物资产上市公司", "condition": "list"},
    {"name": "生产性生物资产国有企业", "condition": "national"},
    {"name": "生产性生物资产明细表", "condition": "all"},
    {"name": "油气资产上市公司", "condition": "list"},
    {"name": "油气资产国有企业", "condition": "national"},
    {"name": "油气资产明细表", "condition": "all"},
    {"name": "使用权资产上市公司", "condition": "list"},
    {"name": "使用权资产国有企业", "condition": "national"},
    {"name": "使用权资产明细表", "condition": "all"},
    {"name": "无形资产国有企业", "condition": "national"},
    {"name": "无形资产上市公司", "condition": "list"},
    {"name": "无形资产明细表", "condition": "all"},
    {"name": "未办妥产权证书的土地使用权情况", "condition": "all"},
    {"name": "未办妥产权证书的无形资产明细表", "condition": "all"},
    {"name": "开发支出", "condition": "all"},
    {"name": "开发支出明细表", "condition": "all"},
    {"name": "商誉账面价值", "condition": "all"},
    {"name": "商誉减值准备", "condition": "all"},
    {"name": "商誉明细表", "condition": "all"},
    {"name": "长期待摊费用", "condition": "all"},
    {"name": "长期待摊费用明细表", "condition": "all"},
    {"name": "未经抵销的递延所得税资产", "condition": "all"},
    {"name": "未经抵销的递延所得税负债", "condition": "all"},
    {"name": "未确认递延所得税资产明细", "condition": "all"},
    {"name": "未确认递延所得税资产的可抵扣亏损将于以下年度到期", "condition": "all"},
    {"name": "其他非流动资产", "condition": "all"},
    {"name": "其他非流动资产明细表", "condition": "all"},
    {"name": "短期借款明细情况", "condition": "all"},
    {"name": "已逾期未偿还的短期借款情况", "condition": "all"},
    {"name": "短期借款明细表", "condition": "all"},
    {"name": "交易性金融负债", "condition":("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "以公允价值计量且其变动计入当期损益的金融负债", "condition": ("oldFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "衍生金融负债", "condition": "all"},
    {"name": "应付票据", "condition": "all"},
    {"name": "应付账款", "condition": "all"},
    {"name": "账龄超过一年的重要应付账款", "condition": "all"},
    {"name": "应付账款明细表", "condition": "all"},
    {"name": "预收款项", "condition": "all"},
    {"name": "预收款项账龄表", "condition": "all"},
    {"name": "账龄一年以上重要的预收款项", "condition": "all"},
    {"name": "预收账款明细表", "condition": "all"},
    {"name": "合同负债", "condition": "newIncomeCriteria"},
    {"name": "应付职工薪酬明细情况", "condition": "all"},
    {"name": "短期薪酬列示", "condition": "all"},
    {"name": "设定提存计划列示", "condition": "all"},
    {"name": "应付职工薪酬明细表", "condition": "all"},
    {"name": "应交税费", "condition": "all"},
    {"name": "应交税费明细表", "condition": "all"},
    {"name": "应交增值税计提", "condition": "all"},
    {"name": "其他应付款汇总", "condition": "all"},
    {"name": "应付利息", "condition": "all"},
    {"name": "重要的已逾期未支付的利息情况", "condition": "all"},
    {"name": "应付利息明细表", "condition": "all"},
    {"name": "应付股利", "condition": "all"},
    {"name": "账龄一年以上重要的应付股利", "condition": "all"},
    {"name": "应付股利明细表", "condition": "all"},
    {"name": "其他应付款项", "condition": "all"},
    {"name": "账龄超过一年的重要其他应付款项", "condition": "all"},
    {"name": "其他应付款明细表", "condition": "all"},
    {"name": "持有待售负债", "condition": "all"},
    {"name": "一年内到期的非流动负债", "condition": "all"},
    {"name": "其他流动负债", "condition": "all"},
    {"name": "短期应付债券", "condition": "all"},
    {"name": "长期借款", "condition": "all"},
    {"name": "长期借款明细表", "condition": "all"},
    {"name": "应付债券", "condition": "all"},
    {"name": "应付债券的增减变动", "condition": "all"},
    {"name": "应付债券明细表", "condition": "all"},
    {"name": "期末发行在外的优先股永续债等金融工具情况", "condition": "all"},
    {"name": "发行在外的优先股永续债等金融工具变动情况", "condition": "all"},
    {"name": "归属于权益工具持有者的信息", "condition": "all"},
    {"name": "租赁负债", "condition": "all"},
    {"name": "租赁负债明细表", "condition": "all"},
    {"name": "长期应付款汇总", "condition": "all"},
    {"name": "长期应付款", "condition": "all"},
    {"name": "长期应付款明细表", "condition": "all"},
    {"name": "专项应付款", "condition": "all"},
    {"name": "专项应付款明细表", "condition": "all"},
    {"name": "长期应付职工薪酬明细情况", "condition": "all"},
    {"name": "设定受益计划义务现值", "condition": "all"},
    {"name": "计划资产", "condition": "all"},
    {"name": "设定受益计划净负债", "condition": "all"},
    {"name": "预计负债", "condition": "all"},
    {"name": "预计负债明细表", "condition": "all"},
    {"name": "递延收益", "condition": "all"},
    {"name": "递延收益中政府补助项目", "condition": "all"},
    {"name": "递延收益明细表", "condition": "all"},
    {"name": "未实现售后回租损益明细表", "condition": "all"},
    {"name": "其他非流动负债", "condition": "all"},
    {"name": "实收资本", "condition": "all"},
    {"name": "股本", "condition": "all"},
    {"name": "其他权益工具", "condition": "all"},
    {"name": "资本公积", "condition": "all"},
    {"name": "其他综合收益", "condition": "all"},
    {"name": "专项储备", "condition": "all"},
    {"name": "盈余公积", "condition": "all"},
    {"name": "未分配利润", "condition": "all"},
    {"name": "营业收入与营业成本", "condition": "all"},
    {"name": "主营业务收入与主营业务成本", "condition": "all"},
    {"name": "主营业务明细表", "condition": "all"},
    {"name": "其他业务收入与其他业务成本", "condition": "all"},
    {"name": "其他业务明细表", "condition": "all"},
    {"name": "税金及附加", "condition": "all"},
    {"name": "销售费用", "condition": "all"},
    {"name": "管理费用", "condition": "all"},
    {"name": "研发费用", "condition": "all"},
    {"name": "财务费用", "condition": "all"},
    {"name": "财务费用分类表", "condition": "all"},
    {"name": "其他收益", "condition": "all"},
    {"name": "投资收益", "condition": "all"},
    {"name": "净敞口套期收益", "condition": "all"},
    {"name": "公允价值变动损益", "condition": "all"},
    {"name": "信用减值损失", "condition": ("newFinancialInstruments","firstTimeNewFinancialInstruments")},
    {"name": "资产减值损失", "condition": "all"},
    {"name": "资产处置收益", "condition": "all"},
    {"name": "营业外收入", "condition": "all"},
    {"name": "营业外支出", "condition": "all"},
    {"name": "所得税费用", "condition": "all"},
    {"name": "会计利润与所得税费用调整过程", "condition": "all"},
    {"name": "可抵扣暂时性差异明细表", "condition": "all"},
    {"name": "应纳税暂时性差异明细表", "condition": "all"},
    {"name": "可抵扣亏损", "condition": "all"},
    {"name": "当期所得税费用计算表", "condition": "all"},
    {"name": "所得税项目计算", "condition": "all"},
    {"name": "收到其他与经营活动有关的现金", "condition": "list"},
    {"name": "支付其他与经营活动有关的现金", "condition": "list"},
    {"name": "收到其他与投资活动有关的现金", "condition": "list"},
    {"name": "支付其他与投资活动有关的现金", "condition": "all"},
    {"name": "收到其他与筹资活动有关的现金", "condition": "list"},
    {"name": "支付其他与筹资活动有关的现金", "condition": "list"},
    {"name": "现金流量表补充资料", "condition": "all"},
    {"name": "现金流补充资料计算", "condition": "all"},
    {"name": "本期支付的取得子公司的现金净额", "condition": "all"},
    {"name": "本期收到的处置子公司的现金净额", "condition": "all"},
    {"name": "现金及现金等价物的构成", "condition": "all"},
    {"name": "所有权或使用权受到限制的资产", "condition": "all"},
    {"name": "外币货币性项目", "condition": "all"},
    {"name": "股份支付总体情况", "condition": "all"},
    {"name": "以权益结算的股份支付情况", "condition": "all"},
    {"name": "以现金结算的股份支付情况", "condition": "all"},
    {"name": "对外担保明细表", "condition": "all"},
    {"name": "其他或有事项明细表", "condition": "all"},
    {"name": "母公司基本情况", "condition": "all"},
    {"name": "其他关联方情况", "condition": "all"},
    {"name": "采购商品接收劳务", "condition": "all"},
    {"name": "出售商品提供劳务", "condition": "all"},
    {"name": "本公司作为出租方", "condition": "all"},
    {"name": "本公司作为承租方", "condition": "all"},
    {"name": "本公司作为承租方当期承担的租赁负债利息支出", "condition": "all"},
    {"name": "本公司作为担保方", "condition": "all"},
    {"name": "本公司作为被担保方", "condition": "all"},
    {"name": "关联方资金拆借", "condition": "all"},
    {"name": "关键管理人员薪酬", "condition": "all"},
    {"name": "应收关联方款项", "condition": "all"},
    {"name": "应付关联方款项", "condition": "all"},
    {"name": "关联方承诺上市公司", "condition": "list"},
    {"name": "非同一控制下企业合并上市公司", "condition": "list"},
    {"name": "合并成本及商誉上市公司", "condition": "list"},
    {"name": "被购买方于购买日可辨认资产负债上市公司", "condition": "list"},
    {"name": "同一控制下企业合并上市公司", "condition": "list"},
    {"name": "合并成本上市公司", "condition": "list"},
    {"name": "合并日被合并方资产负债的账面价值上市公司", "condition": "list"},
    {"name": "单次处置对子公司投资即丧失控制权上市公司", "condition": "list"},
    {"name": "多次处置构成一揽子交易上市公司", "condition": "list"},
    {"name": "多次处置不构成一揽子交易上市公司", "condition": "list"},
    {"name": "其他合并范围增加上市公司", "condition": "list"},
    {"name": " 其他合并范围减少上市公司", "condition": "list"},
    {"name": "企业集团的构成上市公司", "condition": "list"},
    {"name": "重要的非全资子公司上市公司", "condition": "list"},
    {"name": "重要非全资子企业期末资产负债上市公司", "condition": "list"},
    {"name": "重要非全资子企业期初资产负债上市公司", "condition": "list"},
    {"name": "重要非全资子企业本期损益和现金流量情况上市公司", "condition": "list"},
    {"name": "重要非全资子企业上期损益和现金流量情况上市公司", "condition": "list"},
    {"name": "在子公司的所有者权益份额发生变化的情况说明上市公司", "condition": "list"},
    {"name": "交易对于少数股东权益及归属于母公司所有者权益的影响上市公司", "condition": "list"},
    {"name": "重要的合营企业或联营企业上市公司", "condition": "list"},
    {"name": "重要合营企业财务信息本期数上市公司", "condition": "list"},
    {"name": "重要合营企业财务信息上期数上市公司", "condition": "list"},
    {"name": "重要联营企业财务信息本期数上市公司", "condition": "list"},
    {"name": "重要联营企业财务信息上期数上市公司", "condition": "list"},
    {"name": "不重要合营企业和联营企业的汇总信息上市公司", "condition": "list"},
    {"name": "合营企业或联营企业发生的超额亏损上市公司", "condition": "list"},
    {"name": "重要的共同经营上市公司", "condition": "list"},
    {"name": "金融负债按剩余到期日分类期末数上市公司", "condition": "list"},
    {"name": "金融负债按剩余到期日分类期初数上市公司", "condition": "list"},
    {"name": "利率敏感性分析上市公司", "condition": "list"},
    {"name": "外币货币性项目上市公司", "condition": "list"},
    {"name": "外汇风险敏感性分析", "condition": "list"},
    {"name": "以公允价值计量的资产和负债的期末公允价值明细情况上市公司", "condition": "list"},
    {"name": "第三层次公允价值计量项目期初与期末账面价值间的调节信息上市公司", "condition": "list"},
    {"name": "不以公允价值计量的金融资产和金融负债的公允价值情况上市公司", "condition": "list"},
    {"name": "资本承诺上市公司", "condition": "list"},
    {"name": "分部信息业务分部期末数", "condition": "list"},
    {"name": "分部信息业务分部期初数", "condition": "list"},
    {"name": "按收入来源地划分的对外交易收入", "condition": "list"},
    {"name": "按资产所在地划分的非流动资产", "condition": "list"},
    {"name": "本公司对主要客户的依赖程度", "condition": "list"},
    {"name": "非经常性损益上市公司", "condition": "list"},
    {"name": "净资产收益率及每股收益上市公司", "condition": "list"},
    {"name": "净资产收益率计算表", "condition": "list"},
    {"name": "每股收益计算表", "condition": "list"},
    {"name": "分类表", "condition": "all"},
    {"name": "信息分类表", "condition": "all"},
]

def add_hide_sheet():
    path = "D:/auditReport/project/combinetbandnote/nationalmodel.xlsx"
    wb = load_workbook(path)
    sheetnames = wb.sheetnames
    condition = ["all","national","oldFinancialInstruments"]
    for sheetname in sheetnames:
        for table in tables:
            if table["name"]==sheetname:
                if isinstance(table["condition"],str):
                    if table["condition"] in condition:
                        pass
                    else:
                        wb[sheetname].sheet_state ='hidden'
                else:
                    for c in table["condition"]:
                        if c in condition:
                            break
                    wb[sheetname].sheet_state = 'hidden'
    wb.save("nationalmodel.xlsx")


def check_hide_sheet():
    path = "D:/auditReport/project/combinetbandnote/combine1.xlsx"
    wb = load_workbook(path)
    sheetnames = wb.sheetnames
    for sheetname in sheetnames:
        print(wb[sheetname].sheet_state)



if __name__ == '__main__':
    add_hide_sheet()